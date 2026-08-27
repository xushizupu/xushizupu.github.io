#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把可编辑的族谱数据（source.xlsx 或 source.csv）转换为线上页面使用的 data.json。

用法:
    python build.py                # 自动找 source.xlsx，找不到则用 source.csv
    python build.py 我的族谱.xlsx
    python build.py 我的族谱.csv

说明:
    - 源文件第一行必须是表头: 姓名,性别,生,卒,配偶,父亲,字辈,职业,葬于,备注
    - 自动为每条记录生成唯一 id 与 fatherId（父亲指向）
    - 在控制台输出校验警告（父亲找不到、姓名重复、空姓名跳过等）
    - 生成 data.json 后提交到 GitHub 即可自动上线
"""
import csv
import json
import os
import sys
from datetime import date

FIELDS = ["姓名", "性别", "生", "卒", "配偶", "父亲", "字辈", "职业", "葬于", "备注"]


def read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            sys.exit("缺少 openpyxl，请先运行: pip install openpyxl（或改用 .csv 源文件）")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            return [row for row in csv.reader(f) if any(str(c).strip() for c in row)]
    sys.exit("不支持的文件类型: %s（仅支持 .xlsx / .csv）" % ext)


def clean(v):
    return "" if v is None else str(v).strip()


def to_int_or_text(v):
    s = clean(v)
    if not s:
        return ""
    try:
        return int(float(s))
    except ValueError:
        return s


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        for candidate in ("source.xlsx", "source.csv"):
            cand = os.path.join(BASE_DIR, candidate)
            if os.path.exists(cand):
                path = cand
                break
    else:
        if not os.path.isabs(path):
            path = os.path.join(BASE_DIR, path)
    if not path:
        sys.exit("未找到 source.xlsx / source.csv，请把族谱文件放在 build.py 所在目录并命名为 source.xlsx 或 source.csv")

    rows = read_rows(path)
    if not rows:
        sys.exit("源文件为空: %s" % path)

    headers = [clean(h) for h in rows[0]]
    if not headers or headers[0] != "姓名":
        print("警告: 第一列表头不是“姓名”(实际: %r)，请确认表头顺序为 %s" % (headers[0] if headers else "", "、".join(FIELDS)))
    col = {h: i for i, h in enumerate(headers)}
    missing = [f for f in FIELDS if f not in col]
    if missing:
        print("警告: 缺少列: %s（将按空值处理）" % "、".join(missing))

    records = []
    skipped = 0
    for r in rows[1:]:
        def get(field):
            i = col.get(field)
            return clean(r[i]) if (i is not None and i < len(r)) else ""

        name = get("姓名")
        if not name:
            skipped += 1
            continue
        records.append({
            "name": name,
            "gender": get("性别"),
            "birth": to_int_or_text(get("生")),
            "death": to_int_or_text(get("卒")),
            "spouse": get("配偶"),
            "father": get("父亲"),
            "generation": get("字辈"),
            "occupation": get("职业"),
            "burial": get("葬于"),
            "remark": get("备注"),
        })

    if not records:
        sys.exit("没有有效记录（所有行姓名均为空）")

    # 分配唯一 id，并记录同名
    name_ids = {}
    for i, rec in enumerate(records, 1):
        rec["id"] = i
        name_ids.setdefault(rec["name"], []).append(i)

    for name, ids in name_ids.items():
        if len(ids) > 1:
            print("警告: 姓名“%s”出现 %d 次（id %s），fatherId 将指向第一个" % (name, len(ids), "、".join(map(str, ids))))

    # 计算 fatherId
    id_of_first = {name: ids[0] for name, ids in name_ids.items()}
    for rec in records:
        if rec["father"]:
            fid = id_of_first.get(rec["father"], 0)
            if not fid:
                print("警告: id=%d %s 的父亲“%s”在数据中不存在，请检查姓名写法" % (rec["id"], rec["name"], rec["father"]))
            rec["fatherId"] = fid
        else:
            rec["fatherId"] = 0

    # updated 取源文件修改日期：源文件没变时输出完全一致，避免自动构建产生空提交
    src_mtime = date.fromtimestamp(os.path.getmtime(path))
    out = {
        "updated": src_mtime.isoformat(),
        "count": len(records),
        "records": records,
    }
    out_path = os.path.join(BASE_DIR, "data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

    print("完成: %s -> %s（%d 条记录，跳过 %d 行空姓名）" % (path, out_path, len(records), skipped))
    print("请把 data.json 一起提交到 GitHub 即可上线。")


if __name__ == "__main__":
    main()